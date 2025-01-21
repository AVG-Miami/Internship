import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
import os.path
from time import sleep


# from matplotlib import style

def create_and_save_plot(data, ticker, period, stl="classic", filename=None):
    plt.figure(figsize=(10, 6))
    plt.style.use(stl)
    if 'Date' not in data:
        if pd.api.types.is_datetime64_any_dtype(data.index):
            dates = data.index.to_numpy()
            plt.subplot(3, 1, 1)
            plt.plot(dates, data['Close'].values, label='Close Price')
            plt.plot(dates, data['Moving_Average'].values, label='Moving Average')
            plt.title(f"{ticker} Цена акций с течением времени")
            plt.legend()
            plt.subplot(3, 1, 2)
            plt.plot(dates, data['RSI'].values, label='RSI')
            plt.legend()
            plt.subplot(3, 1, 3)
            plt.plot(dates, data['MACD'].values, label='MACD')
            plt.plot(dates, data['MACD_H'].values, label='MACD_H')
            plt.plot(dates, data['MACD_S'].values, label='MACD_S')
            plt.legend()
        else:
            print("Информация о дате отсутствует или не имеет распознаваемого формата.")
            return
    else:
        if not pd.api.types.is_datetime64_any_dtype(data['Date']):
            data['Date'] = pd.to_datetime(data['Date'])

        plt.subplot(3, 1, 1)
        plt.plot(data['Date'], data['Close'], label='Close Price')
        plt.plot(data['Date'], data['Moving_Average'], label='Moving Average')
        plt.subplot(3, 1, 2)
        plt.plot(data['Date'], data['RSI'], label='RSI')
        plt.subplot(3, 1, 3)
        plt.plot(data['Date'], data['MACD'], label='MACD')
        plt.plot(data['Date'], data['MACD_H'], label='MACD_H')
        plt.plot(data['Date'], data['MACD_S'], label='MACD_S')

    plt.xlabel("Дата")
    plt.ylabel("Цена")
    plt.legend()

    if filename is None:
        filename = f"{ticker}_{period}_stock_price_chart.png"

    plt.savefig(filename)
    print(f"График сохранен как {filename}")
    return filename


def export_data_to_exel(data, tikers, graphic):
    """ Экспорт в Exel.
    Экспорт производится в фаил Данные.xlsx
    данные каждого тикера (AAPL, GOOGL и тд) добавляются на отдельной странице,
    если экспортируется тикер, страница которого уже присутсвует в файле, то данная страница перезаписывается
    так же на страницу тикера добавляется график изменения цены сгенерированный ранее, а так же вставляется диаграмма
    цены закрытия
    """
    df = pd.DataFrame(data)
    img = Image(graphic)
    if not (os.path.isfile('Данные.xlsx')):
        excel_writer = pd.ExcelWriter('Данные.xlsx', engine='openpyxl', mode='w')
    else:
        ex_file = load_workbook('Данные.xlsx')
        if tikers in ex_file.sheetnames:  # Если существует страница
            if len(ex_file.sheetnames) > 1:  # и это не едиственная страница в книге
                ex_file.remove(ex_file[tikers])  # То удаляем ее
                ex_file.save("Данные.xlsx")
                excel_writer = pd.ExcelWriter('Данные.xlsx', engine='openpyxl', mode='a')
            else:
                excel_writer = pd.ExcelWriter('Данные.xlsx', engine='openpyxl', mode='w')
        else:
            excel_writer = pd.ExcelWriter('Данные.xlsx', engine='openpyxl', mode='a')

    df.to_excel(excel_writer, sheet_name=f'{tikers}', index=False)
    excel_writer.close()

    #   Вставка картинки графика
    ex_file = load_workbook('Данные.xlsx')
    ws = ex_file[tikers]
    row_num = ws.max_row + 1
    cell_addr = f"A{ws.max_row + 1}"
    size = (200, 200)
    img.anchor = cell_addr
    ws.add_image(img)
    ws.row_dimensions[row_num].height = int(size[1] * .8)
    ws.column_dimensions["A"].width = int(size[0] * .2)
    print('Write to excel file "Данные.xlsx" complete.')
    ex_file.save("Данные.xlsx")

    #   Вставка диаграммы
    ex_file = load_workbook('Данные.xlsx')
    ws = ex_file[tikers]

    values = Reference(worksheet=ws,
                       min_row=2,
                       max_row=64,
                       min_col=4,
                       max_col=4)
    # создаем объект столбчатой диаграммы
    chart = BarChart()
    # добавляем в диаграмму выбранный диапазон значений
    chart.add_data(values, titles_from_data=True)
    # привязываем диаграмму к ячейке `m2`
    ws.add_chart(chart, "m2")
    # определяем размеры диаграммы в сантиметрах
    chart.width = 20
    chart.height = 5
    # сохраняем и смотрим что получилось
    ex_file.save("Данные.xlsx")


def export_data_to_csv(data, f_name):
    """ Экспорт в CSV. """
    df = pd.DataFrame(data)
    df.to_csv(f'{f_name}.csv', index=False)
    print(f'Write to file "{f_name}.csv" complete.')


def my_rg():
    # Создание данных для графиков
    x = [1, 2, 3, 4, 5]
    y1 = [1, 4, 9, 16, 25]
    y2 = [1, 2, 3, 4, 5]

    # Создание первого графика
    plt.subplot(2, 1, 1)  # указываем 2 строки, 1 столбец, выбираем первое место
    plt.plot(x, y1)
    plt.title('График 1')

    # Создание второго графика
    plt.subplot(2, 1, 2)  # указываем 2 строки, 1 столбец, выбираем второе место
    plt.plot(x, y2)
    plt.title('График 2')

    plt.show()
